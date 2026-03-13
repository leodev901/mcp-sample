import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── 공통 스타일 ───────────────────────────────────────────
HEADER_FILL_CALENDAR = PatternFill("solid", fgColor="1F4E79")   # 짙은 파랑 (캘린더)
HEADER_FILL_MAIL     = PatternFill("solid", fgColor="7B2D8B")   # 보라 (메일)
HEADER_FILL_PARAM    = PatternFill("solid", fgColor="833C00")   # 주황-갈색 (파라미터)
HEADER_FONT          = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BODY_FONT            = Font(name="맑은 고딕", size=9)
TITLE_FONT           = Font(name="맑은 고딕", bold=True, size=11)
WRAP_ALIGN           = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN         = Alignment(horizontal="center", vertical="center", wrap_text=True)

BAND_FILL_A = PatternFill("solid", fgColor="EBF3FB")
BAND_FILL_B = PatternFill("solid", fgColor="FFFFFF")

THIN_SIDE   = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

Y_FILL = PatternFill("solid", fgColor="E2EFDA")   # 연두 (활성)
N_FILL = PatternFill("solid", fgColor="FCE4EC")   # 연빨 (비활성)


def set_header_row(ws, cols, fill):
    for col_idx, (text, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font   = HEADER_FONT
        cell.fill   = fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_body_cell(ws, row, col, value, band, center=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = BODY_FONT
    cell.fill      = band
    cell.border    = THIN_BORDER
    cell.alignment = CENTER_ALIGN if center else WRAP_ALIGN
    return cell


# ═══════════════════════════════════════════════════════════════
# Sheet 1 – 도구 목록 (Overview)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "도구 목록(Overview)"
ws1.freeze_panes = "A2"
ws1.row_dimensions[1].height = 36

OV_COLS = [
    ("No",          4),
    ("도구명",      28),
    ("분류",         8),
    ("기능 요약",   40),
    ("반환 타입",   10),
    ("활성화 여부",  9),
    ("LLM 에이전트 호출 조건", 50),
    ("비고",        20),
]
set_header_row(ws1, OV_COLS, HEADER_FILL_CALENDAR)

overview_data = [
    # No  도구명                              분류       기능 요약                                              반환타입    활성  LLM 호출 조건                                              비고
    (1,  "list_calendar_events",             "캘린더",  "MS365 Outlook 캘린더 일정 조회",                     "list[dict]","✅","사용자가 '일정 조회해줘', '캘린더 확인해줘' 요청 시",      "start_date/end_date 필수"),
    (2,  "get_recent_emails",                "메일",    "기간 내 최근 순서로 받은 메일 조회",                 "list[dict]","✅","사용자가 '메일 확인해줘', '오늘 메일 보여줘' 요청 시",     "기본 메일 조회 도구"),
    (3,  "get_unread_emails",                "메일",    "아직 읽지 않은 메일(isRead=false) 조회",             "list[dict]","✅","'안읽은 메일 확인해줘' 요청 시",                          "isRead 필드 포함 반환"),
    (4,  "get_important_or_flagged_emails",  "메일",    "중요도 높거나 플래그 지정된 메일 조회",              "list[dict]","✅","'중요한 메일', '깃발 찍힌 메일' 등 요청 시",              "importance/flag_status 반환"),
    (5,  "search_emails_by_keyword_advanced","메일",    "키워드가 제목/본문에 포함된 메일 풀텍스트 검색",     "list[dict]","✅","'회의 관련 메일 찾아줘' 등 키워드 검색 요청 시",           "ConsistencyLevel:eventual 사용"),
    (6,  "search_emails_by_sender_advanced", "메일",    "특정 발신자의 메일 모아서 조회 (Fallback: $search)", "list[dict]","✅","'존한테 온 메일 모아줘' 등 발신자 지정 요청 시",          "400 오류 시 $search Fallback"),
    (7,  "search_emails_by_attachment",      "메일",    "첨부파일 포함 메일 또는 특정 파일명/확장자 검색",    "list[dict]","✅","'pdf 첨부된 메일', '보고서 파일 있는 메일' 요청 시",        "attachments 필드 포함 반환"),
    (8,  "get_sent_emails",                  "메일",    "보낸편지함(Sent Items) 메일 조회",                   "list[dict]","✅","'내가 보낸 편지들', '발송한 메일' 요청 시",               "sentDateTime 기준 정렬"),
    (9,  "get_email_detail_view",            "메일",    "단일 메일 본문 및 첨부파일 상세 조회",               "dict",      "✅","'첫번째 메일 자세히 읽어줘' 등 상세 조회 요청 시",         "메일 id 필수, 본문 text 형식 반환"),
]

CATEGORY_FILL = {"캘린더": PatternFill("solid", fgColor="DDEBF7"),
                 "메일":   PatternFill("solid", fgColor="F3E6FF")}

for r_idx, row in enumerate(overview_data, start=2):
    band = BAND_FILL_A if r_idx % 2 == 0 else BAND_FILL_B
    ws1.row_dimensions[r_idx].height = 30
    for c_idx, val in enumerate(row, start=1):
        cell = write_body_cell(ws1, r_idx, c_idx, val, band, center=(c_idx in [1,3,5,6]))
        # 분류 컬럼 색칠
        if c_idx == 3:
            cell.fill = CATEGORY_FILL.get(val, band)
        # 활성화 컬럼
        if c_idx == 6:
            cell.fill = Y_FILL if val == "✅" else N_FILL


# ═══════════════════════════════════════════════════════════════
# Sheet 2 – 파라미터 상세 (Params)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("파라미터 상세(Params)")
ws2.freeze_panes = "A2"
ws2.row_dimensions[1].height = 36

PM_COLS = [
    ("No",           4),
    ("도구명",      28),
    ("분류",          8),
    ("파라미터명",   22),
    ("타입",         18),
    ("필수 여부",     9),
    ("기본값",       22),
    ("설명",         50),
    ("비고",         20),
]
set_header_row(ws2, PM_COLS, HEADER_FILL_PARAM)

param_data = [
    # No  도구명                              분류        파라미터명       타입               필수  기본값                              설명                                               비고
    (1,  "list_calendar_events",             "캘린더","start_date",    "str",             "필수","2026-03-01T00:00:00",             "조회 시작일 (ISO 8601)",                           "ISO 8601 형식"),
    (2,  "list_calendar_events",             "캘린더","end_date",      "str",             "필수","2026-03-31T23:59:59",             "조회 종료일 (ISO 8601)",                           "ISO 8601 형식"),
    (3,  "list_calendar_events",             "캘린더","top",           "Optional[int]",   "선택","10",                              "최대 조회 건수",                                   ""),
    (4,  "list_calendar_events",             "캘린더","user_email",    "Optional[str]",   "선택","admin@leodev901.onmicrosoft.com", "조회 대상자 이메일",                               "토큰 사용자 정보로 대체 가능"),
    (5,  "get_recent_emails",                "메일", "tok_k",          "int",             "선택","10",                              "가져올 메일 최대 개수 (1~50)",                     ""),
    (6,  "get_recent_emails",                "메일", "from_date",      "Optional[str]",   "선택","None",                            "조회 시작일 (YYYY-MM-DD)",                         ""),
    (7,  "get_recent_emails",                "메일", "to_date",        "Optional[str]",   "선택","None",                            "조회 종료일 (YYYY-MM-DD)",                         ""),
    (8,  "get_unread_emails",                "메일", "tok_k",          "int",             "선택","10",                              "가져올 안읽은 메일 최대 개수 (1~50)",              ""),
    (9,  "get_unread_emails",                "메일", "from_date",      "Optional[str]",   "선택","None",                            "조회 시작일 (YYYY-MM-DD)",                         ""),
    (10, "get_unread_emails",                "메일", "to_date",        "Optional[str]",   "선택","None",                            "조회 종료일 (YYYY-MM-DD)",                         ""),
    (11, "get_important_or_flagged_emails",  "메일", "tok_k",          "int",             "선택","10",                              "가져올 메일 최대 개수 (1~50)",                     ""),
    (12, "get_important_or_flagged_emails",  "메일", "from_date",      "Optional[str]",   "선택","None",                            "조회 시작일 (YYYY-MM-DD)",                         ""),
    (13, "get_important_or_flagged_emails",  "메일", "to_date",        "Optional[str]",   "선택","None",                            "조회 종료일 (YYYY-MM-DD)",                         ""),
    (14, "get_important_or_flagged_emails",  "메일", "isimportant",    "bool",            "선택","True",                            "중요(high) 메일 포착 여부",                        ""),
    (15, "get_important_or_flagged_emails",  "메일", "isflagged",      "bool",            "선택","True",                            "플래그 지정 메일 포착 여부",                       ""),
    (16, "search_emails_by_keyword_advanced","메일", "keyword",        "str",             "필수","-",                               "검색할 키워드",                                    "필수값"),
    (17, "search_emails_by_keyword_advanced","메일", "tok_k",          "int",             "선택","10",                              "최대 조회 가능 메일 건수",                         ""),
    (18, "search_emails_by_keyword_advanced","메일", "from_date",      "Optional[str]",   "선택","None",                            "조회 일자 제한 시작일 (YYYY-MM-DD)",               ""),
    (19, "search_emails_by_keyword_advanced","메일", "to_date",        "Optional[str]",   "선택","None",                            "조회 일자 제한 종료일 (YYYY-MM-DD)",               ""),
    (20, "search_emails_by_keyword_advanced","메일", "scope",          "Literal[제목/본문/all]","선택","all",                       "검색 대상 필드 (제목, 본문, all 중 택 1)",         "ConsistencyLevel 헤더 필요"),
    (21, "search_emails_by_sender_advanced", "메일", "sender",         "str",             "필수","-",                               "발신자 이메일 주소 또는 이름",                     "필수값. $filter 실패 시 $search Fallback"),
    (22, "search_emails_by_sender_advanced", "메일", "tok_k",          "int",             "선택","10",                              "최대 조회 건수",                                   ""),
    (23, "search_emails_by_sender_advanced", "메일", "from_date",      "Optional[str]",   "선택","None",                            "조회 시작일 (YYYY-MM-DD)",                         ""),
    (24, "search_emails_by_sender_advanced", "메일", "to_date",        "Optional[str]",   "선택","None",                            "조회 종료일 (YYYY-MM-DD)",                         ""),
    (25, "search_emails_by_attachment",      "메일", "tok_k",          "int",             "선택","10",                              "최대 조회 건수",                                   ""),
    (26, "search_emails_by_attachment",      "메일", "from_date",      "Optional[str]",   "선택","None",                            "조회 시작일 (YYYY-MM-DD)",                         ""),
    (27, "search_emails_by_attachment",      "메일", "to_date",        "Optional[str]",   "선택","None",                            "조회 종료일 (YYYY-MM-DD)",                         ""),
    (28, "search_emails_by_attachment",      "메일", "filename",       "Optional[str]",   "선택","None",                            "검색할 첨부파일명 (일부분 일치 허용)",             ""),
    (29, "search_emails_by_attachment",      "메일", "fileext",        "Optional[str]",   "선택","None",                            "검색할 첨부파일 확장자 (예: pdf, docx)",           ""),
    (30, "get_sent_emails",                  "메일", "tok_k",          "int",             "선택","10",                              "가져올 보낸 메일 최대 수",                         ""),
    (31, "get_sent_emails",                  "메일", "from_date",      "Optional[str]",   "선택","None",                            "조회 시작일 (YYYY-MM-DD)",                         "sentDateTime 기준으로 변환"),
    (32, "get_sent_emails",                  "메일", "to_date",        "Optional[str]",   "선택","None",                            "조회 종료일 (YYYY-MM-DD)",                         "sentDateTime 기준으로 변환"),
    (33, "get_email_detail_view",            "메일", "id",             "str",             "필수","-",                               "조회할 메일의 고유 ID",                            "필수값. 다른 도구 결과의 id 값 사용"),
]

MANDATORY_FILL = PatternFill("solid", fgColor="FFF2CC")  # 연노랑 (필수)
OPTIONAL_FILL  = PatternFill("solid", fgColor="F5F5F5")  # 연회색 (선택)

for r_idx, row in enumerate(param_data, start=2):
    band = BAND_FILL_A if r_idx % 2 == 0 else BAND_FILL_B
    ws2.row_dimensions[r_idx].height = 28
    for c_idx, val in enumerate(row, start=1):
        cell = write_body_cell(ws2, r_idx, c_idx, val, band, center=(c_idx in [1, 3, 5, 6]))
        if c_idx == 3:
            cell.fill = CATEGORY_FILL.get(val, band)
        if c_idx == 6:
            cell.fill = MANDATORY_FILL if val == "필수" else OPTIONAL_FILL
            cell.font = Font(name="맑은 고딕", size=9, bold=(val == "필수"))

# ─── 저장 ──────────────────────────────────────────────────────
output_path = "docs/MCP_도구_정의서.xlsx"
import os; os.makedirs("docs", exist_ok=True)
wb.save(output_path)
print(f"OK: {output_path}")
